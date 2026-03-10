"""
Views do Módulo Fiscal — Interface web para monitoramento de NFe/CTe/MDFe/NFSe.
"""

import io
import logging
from datetime import date
from decimal import Decimal

from django.contrib import messages
from django.db.models import Sum, Count, Q
from django.http import JsonResponse, HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.template.loader import render_to_string
from django.urls import reverse
from django.utils import timezone
from django.views import View
from django.views.generic import ListView, DetailView, TemplateView

from .models import ConfiguracaoFiscal, DocumentoFiscalEntrada, LogConsultaSefaz
from .services import (
    enviar_manifestacao, download_nfe_completa, processar_xml_upload,
    EVENTO_CIENCIA, EVENTO_CONFIRMACAO, EVENTO_DESCONHECIMENTO, EVENTO_NAO_REALIZADA,
)

logger = logging.getLogger('fiscal')

MESES = [
    (1, 'Janeiro'), (2, 'Fevereiro'), (3, 'Março'), (4, 'Abril'),
    (5, 'Maio'), (6, 'Junho'), (7, 'Julho'), (8, 'Agosto'),
    (9, 'Setembro'), (10, 'Outubro'), (11, 'Novembro'), (12, 'Dezembro'),
]


def _get_filtro_mes_ano(request):
    """Retorna (mes, ano) a partir dos GET params, com defaults para o mês atual."""
    try:
        ano = int(request.GET.get('ano', date.today().year))
    except (ValueError, TypeError):
        ano = date.today().year
    try:
        mes = int(request.GET.get('mes', date.today().month))
        if mes < 1 or mes > 12:
            mes = date.today().month
    except (ValueError, TypeError):
        mes = date.today().month
    return mes, ano


def _filtrar_por_mes_ano(qs, mes, ano):
    """Filtra queryset de documentos por mês/ano."""
    return qs.filter(
        Q(data_emissao__month=mes, data_emissao__year=ano) |
        Q(data_emissao__isnull=True, criado_em__month=mes, criado_em__year=ano)
    )


def _add_filtro_context(ctx, mes, ano):
    """Adiciona variáveis de mês/ano ao contexto do template."""
    ctx['mes'] = mes
    ctx['ano'] = ano
    ctx['anos_disponiveis'] = range(date.today().year - 5, date.today().year + 1)
    ctx['meses'] = MESES
    return ctx


# ===========================================================================
# Dashboard Fiscal
# ===========================================================================
class FiscalDashboardView(TemplateView):
    template_name = 'fiscal/dashboard.html'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        mes, ano = _get_filtro_mes_ano(self.request)
        _add_filtro_context(ctx, mes, ano)

        qs = DocumentoFiscalEntrada.objects.all()
        qs_periodo = _filtrar_por_mes_ano(qs, mes, ano)

        ctx['total_documentos'] = qs.count()
        ctx['total_periodo'] = qs_periodo.count()
        ctx['valor_total'] = qs_periodo.aggregate(total=Sum('valor_total'))['total'] or Decimal('0')

        # Por status
        ctx['sem_manifestacao'] = qs_periodo.filter(status_manifestacao='sem_manifestacao').count()
        ctx['ciencia'] = qs_periodo.filter(status_manifestacao='ciencia').count()
        ctx['confirmadas'] = qs_periodo.filter(status_manifestacao='confirmada').count()
        ctx['desconhecidas'] = qs_periodo.filter(status_manifestacao='desconhecida').count()
        ctx['nao_realizadas'] = qs_periodo.filter(status_manifestacao='nao_realizada').count()

        # Por tipo
        ctx['total_nfe'] = qs_periodo.filter(tipo_documento='nfe').count()
        ctx['total_cte'] = qs_periodo.filter(tipo_documento='cte').count()
        ctx['total_mdfe'] = qs_periodo.filter(tipo_documento='mdfe').count()
        ctx['total_nfse'] = qs_periodo.filter(tipo_documento='nfse').count()
        ctx['total_cce'] = qs_periodo.filter(tipo_documento='cce').count()
        ctx['total_resumo'] = qs_periodo.filter(tipo_documento='resumo').count()

        # Por papel
        ctx['total_destinatario'] = qs_periodo.filter(papel_empresa='destinatario').count()
        ctx['total_transportador'] = qs_periodo.filter(papel_empresa='transportador').count()

        # Últimos documentos e logs
        ctx['ultimos_documentos'] = qs_periodo.order_by('-criado_em')[:10]
        ctx['ultimos_logs'] = LogConsultaSefaz.objects.order_by('-data_consulta')[:5]
        ctx['config'] = ConfiguracaoFiscal.objects.first()

        return ctx


# ===========================================================================
# Lista de Documentos Fiscais
# ===========================================================================
class DocumentoFiscalListView(ListView):
    model = DocumentoFiscalEntrada
    template_name = 'fiscal/documento_list.html'
    context_object_name = 'documentos'
    paginate_by = 25

    def get_queryset(self):
        qs = super().get_queryset()

        mes, ano = _get_filtro_mes_ano(self.request)
        qs = _filtrar_por_mes_ano(qs, mes, ano)

        status = self.request.GET.get('status')
        if status and status != 'todos':
            qs = qs.filter(status_manifestacao=status)

        tipo = self.request.GET.get('tipo')
        if tipo and tipo != 'todos':
            qs = qs.filter(tipo_documento=tipo)

        situacao = self.request.GET.get('situacao')
        if situacao and situacao != 'todos':
            qs = qs.filter(situacao=situacao)

        papel = self.request.GET.get('papel')
        if papel and papel != 'todos':
            qs = qs.filter(papel_empresa=papel)

        busca = self.request.GET.get('q', '').strip()
        if busca:
            qs = qs.filter(
                Q(emitente_nome__icontains=busca) |
                Q(emitente_cnpj__icontains=busca) |
                Q(chave_acesso__icontains=busca) |
                Q(numero_nf__icontains=busca)
            )

        return qs

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)

        mes, ano = _get_filtro_mes_ano(self.request)
        _add_filtro_context(ctx, mes, ano)

        qs_mes = _filtrar_por_mes_ano(DocumentoFiscalEntrada.objects.all(), mes, ano)
        ctx['total_geral'] = qs_mes.count()
        ctx['total_sem_manifestacao'] = qs_mes.filter(status_manifestacao='sem_manifestacao').count()
        ctx['total_ciencia'] = qs_mes.filter(status_manifestacao='ciencia').count()
        ctx['total_confirmadas'] = qs_mes.filter(status_manifestacao='confirmada').count()
        ctx['valor_total'] = qs_mes.aggregate(t=Sum('valor_total'))['t'] or Decimal('0')

        ctx['filtro_status'] = self.request.GET.get('status', 'todos')
        ctx['filtro_tipo'] = self.request.GET.get('tipo', 'todos')
        ctx['filtro_situacao'] = self.request.GET.get('situacao', 'todos')
        ctx['filtro_papel'] = self.request.GET.get('papel', 'todos')
        ctx['filtro_busca'] = self.request.GET.get('q', '')

        return ctx


# ===========================================================================
# Detalhe do Documento
# ===========================================================================
class DocumentoFiscalDetailView(DetailView):
    model = DocumentoFiscalEntrada
    template_name = 'fiscal/documento_detail.html'
    context_object_name = 'doc'

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx['config'] = ConfiguracaoFiscal.objects.first()

        doc = self.object

        # CC-e: documento original
        if doc.tipo_documento == 'cce' and doc.chave_referencia:
            ctx['doc_original'] = DocumentoFiscalEntrada.objects.filter(
                chave_acesso=doc.chave_referencia
            ).first()

        # CC-e vinculadas
        ctx['cces_vinculadas'] = DocumentoFiscalEntrada.objects.filter(
            chave_referencia=doc.chave_acesso,
            tipo_documento='cce'
        ).order_by('-data_emissao')

        return ctx


# ===========================================================================
# Ações de Manifestação
# ===========================================================================
class ManifestacaoView(View):
    """Envia manifestação do destinatário à SEFAZ."""

    def post(self, request, pk):
        doc = get_object_or_404(DocumentoFiscalEntrada, pk=pk)
        config = ConfiguracaoFiscal.objects.filter(ativo=True).first()

        if not config:
            messages.error(request, 'Nenhuma configuração fiscal ativa.')
            return redirect('fiscal:documento_detail', pk=pk)

        tipo_evento = request.POST.get('tipo_evento')
        justificativa = request.POST.get('justificativa', '')

        evento_map = {
            'ciencia': EVENTO_CIENCIA,
            'confirmada': EVENTO_CONFIRMACAO,
            'desconhecida': EVENTO_DESCONHECIMENTO,
            'nao_realizada': EVENTO_NAO_REALIZADA,
        }

        codigo_evento = evento_map.get(tipo_evento)
        if not codigo_evento:
            messages.error(request, 'Tipo de evento inválido.')
            return redirect('fiscal:documento_detail', pk=pk)

        if tipo_evento in ('desconhecida', 'nao_realizada') and not justificativa.strip():
            messages.error(request, 'Justificativa é obrigatória para este tipo de manifestação.')
            return redirect('fiscal:documento_detail', pk=pk)

        sucesso, mensagem = enviar_manifestacao(config, doc, codigo_evento, justificativa)

        if sucesso:
            messages.success(request, f'Manifestação registrada com sucesso! {mensagem}')
        else:
            messages.error(request, f'Erro ao enviar manifestação: {mensagem}')

        return redirect('fiscal:documento_detail', pk=pk)


# ===========================================================================
# Download XML Completo
# ===========================================================================
class DownloadXMLView(View):
    """Baixa o XML completo da NFe após Ciência da Operação."""

    def get(self, request, pk):
        doc = get_object_or_404(DocumentoFiscalEntrada, pk=pk)

        if doc.xml_completo:
            response = HttpResponse(doc.xml_completo, content_type='application/xml')
            response['Content-Disposition'] = (
                f'attachment; filename="{doc.get_tipo_documento_display()}_{doc.chave_acesso}.xml"'
            )
            return response

        config = ConfiguracaoFiscal.objects.filter(ativo=True).first()
        if not config:
            messages.error(request, 'Sem configuração fiscal.')
            return redirect('fiscal:documento_detail', pk=pk)

        xml_bytes, mensagem = download_nfe_completa(config, doc.chave_acesso)
        if xml_bytes:
            doc.xml_completo = xml_bytes.decode('utf-8', errors='ignore')
            doc.save(update_fields=['xml_completo'])

            response = HttpResponse(xml_bytes, content_type='application/xml')
            response['Content-Disposition'] = (
                f'attachment; filename="{doc.get_tipo_documento_display()}_{doc.chave_acesso}.xml"'
            )
            return response
        else:
            messages.warning(request, f'Não foi possível baixar o XML: {mensagem}')
            return redirect('fiscal:documento_detail', pk=pk)


# ===========================================================================
# Forçar Consulta Manual
# ===========================================================================
class ForcarConsultaView(View):
    """Dispara manualmente a tarefa de consulta à SEFAZ."""

    def post(self, request):
        from .tasks import buscar_documentos_fiscais, buscar_nfse
        buscar_documentos_fiscais.delay()
        buscar_nfse.delay()
        messages.success(request, 'Consulta à SEFAZ disparada. Os documentos aparecerão em instantes.')
        return redirect('fiscal:dashboard')


# ===========================================================================
# Importar XML manualmente
# ===========================================================================
class ImportarXMLView(View):
    """Upload manual de XMLs de NFe/CTe/MDFe/NFSe/CCe."""

    def get(self, request):
        return render(request, 'fiscal/importar_xml.html')

    def post(self, request):
        arquivos = request.FILES.getlist('arquivos_xml')
        if not arquivos:
            messages.error(request, 'Nenhum arquivo selecionado.')
            return redirect('fiscal:importar_xml')

        config = ConfiguracaoFiscal.objects.first()
        cnpj = config.cnpj if config else ''

        sucesso = 0
        erros = []

        for arquivo in arquivos:
            if not arquivo.name.lower().endswith('.xml'):
                erros.append(f'{arquivo.name}: não é um arquivo XML')
                continue

            try:
                xml_content = arquivo.read()
                dados, erro = processar_xml_upload(xml_content, cnpj)
                if erro:
                    erros.append(f'{arquivo.name}: {erro}')
                    continue

                chave = dados.get('chave_acesso', '')
                if not chave:
                    erros.append(f'{arquivo.name}: chave de acesso não encontrada no XML')
                    continue

                obj, created = DocumentoFiscalEntrada.objects.update_or_create(
                    chave_acesso=chave,
                    defaults={
                        'tipo_documento': dados.get('tipo_documento', 'nfe'),
                        'papel_empresa': dados.get('papel_empresa', 'destinatario'),
                        'origem': 'upload',
                        'emitente_cnpj': dados.get('emitente_cnpj', ''),
                        'emitente_nome': dados.get('emitente_nome', ''),
                        'emitente_ie': dados.get('emitente_ie', ''),
                        'numero_nf': dados.get('numero_nf', ''),
                        'serie': dados.get('serie', ''),
                        'data_emissao': dados.get('data_emissao'),
                        'valor_total': dados.get('valor_total'),
                        'natureza_operacao': dados.get('natureza_operacao', '')[:255],
                        'xml_completo': dados.get('xml_completo', ''),
                        'chave_referencia': dados.get('chave_referencia', ''),
                        'texto_correcao': dados.get('texto_correcao', ''),
                        'criado_por': request.user if request.user.is_authenticated else None,
                    }
                )

                # Salvar arquivo
                arquivo.seek(0)
                obj.arquivo_xml.save(arquivo.name, arquivo, save=True)
                sucesso += 1

            except Exception as e:
                erros.append(f'{arquivo.name}: {str(e)}')
                logger.error(f'Erro ao importar XML {arquivo.name}: {e}', exc_info=True)

        if sucesso:
            messages.success(request, f'{sucesso} documento(s) importado(s) com sucesso!')
        if erros:
            for erro in erros[:5]:
                messages.warning(request, erro)
            if len(erros) > 5:
                messages.warning(request, f'...e mais {len(erros) - 5} erro(s).')

        return redirect('fiscal:documento_list')


# ===========================================================================
# Exportar Excel
# ===========================================================================
class ExportarExcelView(View):
    """Exporta documentos fiscais filtrados para Excel."""

    def get(self, request):
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter

        mes, ano = _get_filtro_mes_ano(request)
        qs = _filtrar_por_mes_ano(DocumentoFiscalEntrada.objects.all(), mes, ano)

        status = request.GET.get('status')
        if status and status != 'todos':
            qs = qs.filter(status_manifestacao=status)
        tipo = request.GET.get('tipo')
        if tipo and tipo != 'todos':
            qs = qs.filter(tipo_documento=tipo)
        situacao = request.GET.get('situacao')
        if situacao and situacao != 'todos':
            qs = qs.filter(situacao=situacao)
        papel = request.GET.get('papel')
        if papel and papel != 'todos':
            qs = qs.filter(papel_empresa=papel)
        busca = request.GET.get('q', '').strip()
        if busca:
            qs = qs.filter(
                Q(emitente_nome__icontains=busca) |
                Q(emitente_cnpj__icontains=busca) |
                Q(chave_acesso__icontains=busca) |
                Q(numero_nf__icontains=busca)
            )

        wb = Workbook()
        ws = wb.active
        ws.title = "Documentos Fiscais"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        currency_format = 'R$ #,##0.00'

        headers = [
            'Tipo', 'Papel', 'Nº NF', 'Série', 'Emitente', 'CNPJ Emitente',
            'Valor (R$)', 'Emissão', 'Situação', 'Manifestação', 'Origem',
            'Chave de Acesso', 'Observações',
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        for row_num, doc in enumerate(qs, 2):
            data = [
                doc.get_tipo_documento_display(),
                doc.get_papel_empresa_display(),
                doc.numero_nf or '',
                doc.serie or '',
                doc.emitente_nome or '',
                doc.cnpj_formatado,
                float(doc.valor_total) if doc.valor_total else '',
                doc.data_emissao.strftime('%d/%m/%Y') if doc.data_emissao else '',
                doc.situacao.capitalize() if doc.situacao else '',
                doc.get_status_manifestacao_display(),
                doc.get_origem_display(),
                doc.chave_acesso,
                doc.observacoes or '',
            ]
            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = thin_border
                if col == 7 and isinstance(value, float):
                    cell.number_format = currency_format

        column_widths = [10, 16, 10, 6, 35, 20, 15, 12, 12, 22, 14, 46, 30]
        for i, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = 'A2'

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        mes_nome = dict(MESES).get(mes, '')
        filename = f"documentos_fiscais_{mes_nome}_{ano}_{timezone.now().strftime('%Y%m%d_%H%M')}.xlsx"
        response = HttpResponse(
            buffer.getvalue(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ===========================================================================
# Exportar PDF (via weasyprint)
# ===========================================================================
class ExportarPDFView(View):
    """Exporta documentos fiscais filtrados para PDF via weasyprint."""

    def get(self, request):
        try:
            import weasyprint
        except ImportError:
            messages.error(request, 'Biblioteca weasyprint não instalada. Use a exportação Excel.')
            return redirect('fiscal:documento_list')

        from django.conf import settings

        mes, ano = _get_filtro_mes_ano(request)
        qs = _filtrar_por_mes_ano(DocumentoFiscalEntrada.objects.all(), mes, ano)

        status = request.GET.get('status')
        if status and status != 'todos':
            qs = qs.filter(status_manifestacao=status)
        tipo = request.GET.get('tipo')
        if tipo and tipo != 'todos':
            qs = qs.filter(tipo_documento=tipo)
        papel = request.GET.get('papel')
        if papel and papel != 'todos':
            qs = qs.filter(papel_empresa=papel)

        documentos = qs
        valor_total = qs.aggregate(t=Sum('valor_total'))['t'] or Decimal('0')
        mes_nome = dict(MESES).get(mes, '')

        tipo_counts = qs.values('tipo_documento').annotate(c=Count('id'))
        tipo_map = {item['tipo_documento']: item['c'] for item in tipo_counts}

        html = render_to_string('fiscal/export_pdf.html', {
            'documentos': documentos,
            'mes': mes,
            'ano': ano,
            'mes_nome': mes_nome,
            'valor_total': valor_total,
            'total_docs': qs.count(),
            'total_nfe': tipo_map.get('nfe', 0),
            'total_cte': tipo_map.get('cte', 0),
            'total_mdfe': tipo_map.get('mdfe', 0),
            'data_geracao': timezone.now(),
        })

        pdf = weasyprint.HTML(string=html, base_url=str(settings.BASE_DIR)).write_pdf()

        filename = f"documentos_fiscais_{mes_nome}_{ano}.pdf"
        response = HttpResponse(pdf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response


# ===========================================================================
# Configuração Fiscal
# ===========================================================================
class ConfiguracaoFiscalView(View):
    """Tela de configuração do módulo fiscal (certificado, CNPJ, senha)."""

    def get(self, request):
        config = ConfiguracaoFiscal.objects.first()
        logs = LogConsultaSefaz.objects.order_by('-data_consulta')[:20]
        return render(request, 'fiscal/configuracao.html', {
            'config': config,
            'logs': logs,
        })

    def post(self, request):
        config, _ = ConfiguracaoFiscal.objects.get_or_create(pk=1)

        config.cnpj = request.POST.get('cnpj', config.cnpj)
        config.uf_codigo = request.POST.get('uf_codigo', config.uf_codigo)
        config.ambiente = int(request.POST.get('ambiente', config.ambiente))
        config.caminho_certificado = request.POST.get('caminho_certificado', config.caminho_certificado)
        config.ativo = request.POST.get('ativo') == 'on'

        nova_senha = request.POST.get('senha_certificado', '').strip()
        if nova_senha:
            config.senha_certificado = nova_senha

        config.save()
        messages.success(request, 'Configuração fiscal salva com sucesso!')

        logs = LogConsultaSefaz.objects.order_by('-data_consulta')[:20]
        return render(request, 'fiscal/configuracao.html', {
            'config': config,
            'logs': logs,
        })


# ===========================================================================
# API: Totais (para refresh via AJAX/HTMX)
# ===========================================================================
class FiscalTotaisView(View):
    def get(self, request):
        mes, ano = _get_filtro_mes_ano(request)
        qs = _filtrar_por_mes_ano(DocumentoFiscalEntrada.objects.all(), mes, ano)
        data = {
            'total': qs.count(),
            'sem_manifestacao': qs.filter(status_manifestacao='sem_manifestacao').count(),
            'ciencia': qs.filter(status_manifestacao='ciencia').count(),
            'confirmadas': qs.filter(status_manifestacao='confirmada').count(),
            'valor_total': str(qs.aggregate(t=Sum('valor_total'))['t'] or '0'),
        }
        return JsonResponse(data)
